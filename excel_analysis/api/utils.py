import openpyxl
import tempfile
import os
from typing import List, Dict, Any


class ExcelProcessor:
    """
    More advanced Excel processing using openpyxl, it was required to handle more complex files like example provided in task description:
    https://drive.google.com/file/d/1QVab0xApzY26CIn0zN_ISqRCE3P6UpxJ/view?usp=sharing
    - No header in the first row
    - Column names with extra spaces
    - Numeric values stored as text with commas
    - Multiple sheets
    etc.
    """

    @staticmethod
    def find_best_header_row(worksheet) -> int:
        best_row, best_count = 1, 0
        for row in range(1, min(11, worksheet.max_row + 1)):
            text_cells = sum(
                1
                for col in range(1, worksheet.max_column + 1)
                if isinstance(worksheet.cell(row=row, column=col).value, str)
            )
            if text_cells > best_count:
                best_row, best_count = row, text_cells
        return best_row

    @staticmethod
    def find_column_matches(
        headers: List[str], target_columns: List[str]
    ) -> Dict[str, int]:
        matches = {}
        for target in target_columns:
            for idx, header in enumerate(headers):
                if header and target.lower() in header.lower():
                    matches[target] = idx
                    break
        return matches

    @staticmethod
    def extract_numeric_values(worksheet, header_row: int, col_idx: int) -> List[float]:
        values = []
        for row in range(header_row + 1, worksheet.max_row + 1):
            val = worksheet.cell(row=row, column=col_idx + 1).value
            if isinstance(val, (int, float)):
                values.append(float(val))
            elif isinstance(val, str):
                try:
                    values.append(float(val.replace(",", "").replace("$", "").strip()))
                except ValueError:
                    pass
        return values

    @staticmethod
    def find_columns_in_excel(
        file_path: str, target_columns: List[str], sheet_name: str = None
    ) -> List[Dict[str, Any]]:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active

        header_row = ExcelProcessor.find_best_header_row(sheet)
        headers = [
            str(sheet.cell(row=header_row, column=c).value or "").strip()
            for c in range(1, sheet.max_column + 1)
        ]

        matches = ExcelProcessor.find_column_matches(headers, target_columns)

        summary = []
        for name, idx in matches.items():
            values = ExcelProcessor.extract_numeric_values(sheet, header_row, idx)
            if values:
                total = sum(values)
                avg = total / len(values)
                summary.append(
                    {"column": name, "sum": round(total, 2), "avg": round(avg, 2)}
                )

        workbook.close()
        return summary


def process_excel_file(
    file, columns: List[str], sheet_name: str = None
) -> List[Dict[str, Any]]:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        for chunk in file.chunks():
            tmp.write(chunk)
        path = tmp.name

    try:
        return ExcelProcessor.find_columns_in_excel(path, columns, sheet_name)
    finally:
        if os.path.exists(path):
            os.remove(path)
