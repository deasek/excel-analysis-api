# api/tests.py

import os
import tempfile
import openpyxl
from django.test import TestCase, Client
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from rest_framework import status
import json


class ExcelAnalysisAPITestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.analyze_url = reverse("analyze_excel")

        self.simple_excel_file = self._create_simple_excel_file()
        self.complex_excel_file = self._create_complex_excel_file()
        self.invalid_file = self._create_text_file()

    def tearDown(self):
        for file_path in [
            self.simple_excel_file,
            self.complex_excel_file,
            self.invalid_file,
        ]:
            if os.path.exists(file_path):
                os.unlink(file_path)

    def _create_simple_excel_file(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # headers
        worksheet["A1"] = "Product"
        worksheet["B1"] = "Price"
        worksheet["C1"] = "Quantity"

        # data
        data = [
            ["Item 1", 100.50, 10],
            ["Item 2", 200.00, 5],
            ["Item 3", 150.75, 8],
            ["Item 4", 75.25, 15],
        ]

        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        workbook.save(temp_file.name)
        workbook.close()
        return temp_file.name

    def _create_complex_excel_file(self):
        """Create a complex Excel file with multiple header rows"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Add title row (should be ignored)
        worksheet["A1"] = "*** PRICE LIST 2024 ***"

        # Add empty row
        # Row 2 is empty

        # Add actual headers (row 3)
        worksheet["A3"] = "ID"
        worksheet["B3"] = "Description"
        worksheet["C3"] = " CURRENT USD"  # Note the space
        worksheet["D3"] = " CURRENT CAD"  # Note the space
        worksheet["E3"] = "THIS IS THE NEW PRICE IN USD"  # Long header

        # Add section header (row 4)
        worksheet["A4"] = "ELECTRONICS SECTION"

        # Add data starting from row 5
        data = [
            ["E001", "Laptop", 1200.00, 1440.00, 1250.00],
            ["E002", "Mouse", 25.50, 30.60, 27.00],
            ["E003", "Keyboard", 85.75, 102.90, 89.00],
            ["E004", "Monitor", 350.00, 420.00, 365.00],
        ]

        for row_idx, row_data in enumerate(data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        workbook.save(temp_file.name)
        workbook.close()
        return temp_file.name

    def _create_text_file(self):
        """Create a text file (invalid format)"""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".txt", mode="w")
        temp_file.write("This is not an Excel file")
        temp_file.close()
        return temp_file.name


class ExcelAnalysisTestCase(ExcelAnalysisAPITestCase):
    """Test Excel analysis endpoint"""

    def test_analyze_simple_excel_success(self):
        """Test successful analysis of simple Excel file"""
        with open(self.simple_excel_file, "rb") as f:
            file_data = SimpleUploadedFile(name="test.xlsx", content=f.read())

        response = self.client.post(
            self.analyze_url, {"file": file_data, "columns": ["price", "quantity"]}
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()
        self.assertIn("file", data)
        self.assertIn("summary", data)
        self.assertEqual(data["file"], "test.xlsx")

        # Check summary structure
        summary = data["summary"]
        self.assertIsInstance(summary, list)
        self.assertGreater(len(summary), 0)

        # Check if price and quantity columns were found
        columns_found = [item["column"] for item in summary]
        self.assertIn("price", columns_found)
        self.assertIn("quantity", columns_found)

        # Check data types and values
        for item in summary:
            self.assertIn("column", item)
            self.assertIn("sum", item)
            self.assertIn("avg", item)
            self.assertIsInstance(item["sum"], (int, float))
            self.assertIsInstance(item["avg"], (int, float))

    def test_analyze_complex_excel_success(self):
        """Test successful analysis of complex Excel file with smart header detection"""
        with open(self.complex_excel_file, "rb") as f:
            file_data = SimpleUploadedFile(
                name="complex.xlsx",
                content=f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        response = self.client.post(
            self.analyze_url,
            {"file": file_data, "columns": ["current usd", "current cad", "price"]},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()
        summary = data["summary"]

        columns_found = [item["column"] for item in summary]
        self.assertIn("current usd", columns_found)
        self.assertIn("current cad", columns_found)

        for item in summary:
            if item["column"] == "current usd":
                self.assertGreater(item["sum"], 1000)  # Should be 1661.25
                self.assertGreater(item["avg"], 400)  # Should be 415.31

    def test_analyze_fuzzy_column_matching(self):
        """Test fuzzy column matching with short names"""
        with open(self.complex_excel_file, "rb") as f:
            file_data = SimpleUploadedFile(
                name="complex.xlsx",
                content=f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        response = self.client.post(
            self.analyze_url,
            {
                "file": file_data,
                "columns": ["usd", "cad"],  # Short forms should still match
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()
        summary = data["summary"]
        columns_found = [item["column"] for item in summary]

        self.assertIn("usd", columns_found)
        self.assertIn("cad", columns_found)

    def test_no_file_uploaded(self):
        """Test error when no file is uploaded"""
        response = self.client.post(self.analyze_url, {"columns": ["price"]})

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

        data = response.json()
        self.assertIn("error", data)
        self.assertIn("details", data)
        self.assertIn("file", data["details"])

    def test_invalid_file_type(self):
        """Test error when uploading non-Excel file"""
        with open(self.invalid_file, "rb") as f:
            file_data = SimpleUploadedFile(
                name="test.txt", content=f.read(), content_type="text/plain"
            )

        response = self.client.post(
            self.analyze_url, {"file": file_data, "columns": ["price"]}
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

        data = response.json()
        self.assertIn("error", data)
        self.assertIn("details", data)
        self.assertIn("file", data["details"])

    def test_no_columns_specified(self):
        """Test error when no columns are specified"""
        with open(self.simple_excel_file, "rb") as f:
            file_data = SimpleUploadedFile(
                name="test.xlsx",
                content=f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        response = self.client.post(
            self.analyze_url,
            {
                "file": file_data
                # No columns specified
            },
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

        data = response.json()
        self.assertIn("error", data)
        self.assertIn("details", data)
        self.assertIn("columns", data["details"])

    def test_empty_columns_list(self):
        """Test error when empty columns list is provided"""
        with open(self.simple_excel_file, "rb") as f:
            file_data = SimpleUploadedFile(
                name="test.xlsx",
                content=f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        response = self.client.post(
            self.analyze_url, {"file": file_data, "columns": []}
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_nonexistent_columns(self):
        """Test behavior when requested columns don't exist in file"""
        with open(self.simple_excel_file, "rb") as f:
            file_data = SimpleUploadedFile(
                name="test.xlsx",
                content=f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        response = self.client.post(
            self.analyze_url,
            {
                "file": file_data,
                "columns": ["nonexistent_column", "another_fake_column"],
            },
        )

        # Should return 200 but with empty summary
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()
        self.assertIn("summary", data)
        self.assertEqual(len(data["summary"]), 0)  # No columns found

    def test_mixed_valid_invalid_columns(self):
        """Test with mix of valid and invalid column names"""
        with open(self.simple_excel_file, "rb") as f:
            file_data = SimpleUploadedFile(
                name="test.xlsx",
                content=f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        response = self.client.post(
            self.analyze_url,
            {"file": file_data, "columns": ["price", "nonexistent_column", "quantity"]},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        data = response.json()
        summary = data["summary"]
        columns_found = [item["column"] for item in summary]

        # Should only find the valid columns
        self.assertIn("price", columns_found)
        self.assertIn("quantity", columns_found)
        self.assertNotIn("nonexistent_column", columns_found)


class ExcelProcessorTestCase(TestCase):
    """Test the ExcelProcessor utility class directly"""

    def setUp(self):
        from api.utils import ExcelProcessor

        self.processor = ExcelProcessor()

    def test_fuzzy_column_matching(self):
        from api.utils import ExcelProcessor

        headers = ["ID", "Product Name", " CURRENT USD", " CURRENT CAD", "Description"]
        target_columns = ["current usd", "current cad", "usd", "product"]

        matches = ExcelProcessor.find_column_matches(headers, target_columns)

        self.assertIn("current usd", matches)
        self.assertIn("current cad", matches)
        self.assertEqual(matches["current usd"], 2)
        self.assertEqual(matches["current cad"], 3)
